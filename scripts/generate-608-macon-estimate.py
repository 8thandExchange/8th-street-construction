"""
Generate corrected Habitat estimate for 608 Macon Ave from permit set A2.2.
Run: python3 scripts/generate-608-macon-estimate.py
"""
from __future__ import annotations

import math
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data/estimates/608-macon-habitat-estimate-corrected.xlsx"
DOWNLOADS = Path.home() / "Downloads/608-Macon-Habitat-Estimate-CORRECTED.xlsx"

# ── Authoritative areas from permit set A2.2 (Booker + Vick, May 2026) ──
HEATED_SF = 1425
FIRST_FLOOR_SF = 952
SECOND_FLOOR_SF = 473
GARAGE_SF = 252
PORCH_SF = 260
SLAB_SF = FIRST_FLOOR_SF + GARAGE_SF + 120  # first + garage + porch slab allowance
ROOF_ASPHALT_SF = 1580  # main roof incl. overhangs
ROOF_METAL_SF = PORCH_SF  # porch awnings only per A2.3
WALL_SFG = 2900  # exterior wall area (siding)
DRYWALL_SF = 9200  # hang + finish surface area

WASTE = 0.1
OH_RATE = 0.08  # Habitat program; lower than commercial 20%
CONTINGENCY = 0.05
DONATION_CREDIT = 12000  # appliances, fixtures, misc donated materials

# Division labor/material scalars vs bare market — volunteer program + nonprofit purchasing
HABITAT_DIV_SCALE: dict[str, tuple[float, float]] = {
    "DIV-01": (0.92, 0.58),
    "DIV-03": (0.90, 0.82),
    "DIV-06": (0.84, 0.42),
    "DIV-07": (0.86, 0.50),
    "DIV-08": (0.92, 0.65),
    "DIV-09": (0.86, 0.38),
    "DIV-10": (0.88, 0.65),
    "DIV-11": (0.30, 0.50),
    "DIV-12": (0.88, 0.48),
    "DIV-22": (0.90, 0.80),
    "DIV-23": (0.92, 0.84),
    "DIV-26": (0.90, 0.82),
    "DIV-31": (0.88, 0.78),
    "DIV-32": (0.86, 0.70),
    "DIV-33": (1.00, 0.88),
}


@dataclass
class Line:
    div: str
    desc: str
    qty: float
    unit: str
    mat: float  # market-rate material unit
    lab: float  # market-rate labor unit
    notes: str = ""


def habitat_mat(item: Line) -> float:
    sm, _ = HABITAT_DIV_SCALE.get(item.div, (1.0, 1.0))
    return round(item.mat * sm, 2)


def habitat_lab(item: Line) -> float:
    _, sl = HABITAT_DIV_SCALE.get(item.div, (1.0, 1.0))
    return round(item.lab * sl, 2)


def q(qty: float) -> float:
    return round(qty * (1 + WASTE), 2)


def lines() -> list[Line]:
    cy_slab = SLAB_SF * (4 / 12) / 27
    cy_fill = SLAB_SF * (4 / 12) / 27
    thickened_edge_lf = 248  # perimeter thickened slab edge, not strip footings

    return [
        # DIV-01 GENERAL
        Line("DIV-01", "Building permit + plan review (Augusta-Richmond)", 1, "LS", 2800, 400,
             "CityView; residential new construction"),
        Line("DIV-01", "Trade permits (elec/plumb/mech — sub-pulled)", 1, "LS", 450, 150, ""),
        Line("DIV-01", "Site supervision / GC management (22 wks)", 22, "WK", 0, 650,
             "Habitat program; partial staff allocation"),
        Line("DIV-01", "Mobilization + site signage + erosion control setup", 1, "LS", 850, 1150, ""),
        Line("DIV-01", "Temporary facilities (porta-john + dumpster, 6 mo)", 6, "MO", 420, 180, ""),
        Line("DIV-01", "Final cleaning prior to CO", 1, "LS", 180, 620, ""),
        Line("DIV-01", "Project insurance / builder's risk (allowance)", 1, "LS", 2200, 0, ""),
        Line("DIV-01", "Boundary survey (if not in soft costs)", 1, "LS", 2400, 0, "Phase 1 deliverable"),
        Line("DIV-01", "Truss / framing package engineering", 1, "LS", 2800, 400, "Wood supplier"),
        # DIV-03 CONCRETE — slab-on-grade w/ thickened edges (NOT 15x14 strip footings)
        Line("DIV-03", f"4\" SOG concrete ({SLAB_SF} SF)", cy_slab, "CY", 168, 92,
             "Permit A2.1: slab w/ WWF, broom finish"),
        Line("DIV-03", "6x6 W1.4/W1.4 WWF", SLAB_SF, "SF", 0.38, 0.12, ""),
        Line("DIV-03", "6-mil vapor barrier under slab", SLAB_SF, "SF", 0.18, 0.08, ""),
        Line("DIV-03", "4\" granular fill under slab", cy_fill, "CY", 42, 28, ""),
        Line("DIV-03", "Thickened slab edge / turn-down (per A2.1)", thickened_edge_lf, "LF", 18, 22,
             "Replaces estimator's 15\"x14\" strip footing — wrong type"),
        Line("DIV-03", "Control / expansion joints", 180, "LF", 2.5, 4.5, ""),
        Line("DIV-03", "Simpson post base + 8x8 porch columns", 4, "EA", 85, 125, "Front porch per elev."),
        Line("DIV-03", "Termite pre-treatment (GA letter)", 1, "LS", 450, 150, ""),
        # DIV-06 FRAMING
        Line("DIV-06", "Prefabricated roof trusses (supply + set)", ROOF_ASPHALT_SF, "SF", 2.85, 1.45,
             "Pro set; volunteers stage only"),
        Line("DIV-06", "Wall framing lumber package (2x4 @ 16\" OC)", HEATED_SF, "SF", 4.25, 2.15,
             "Includes plates, headers, blocking"),
        Line("DIV-06", "Wood I-joist second floor system (473 SF)", SECOND_FLOOR_SF, "SF", 5.8, 3.2,
             "Per A2.3 wall section — NOT 2x4 joists"),
        Line("DIV-06", "3/4\" T&G floor sheathing (2nd floor)", SECOND_FLOOR_SF, "SF", 1.35, 0.85, ""),
        Line("DIV-06", "Roof sheathing 7/16\" OSB", ROOF_ASPHALT_SF, "SF", 0.72, 0.42, ""),
        Line("DIV-06", "5/8\" ext. sheathing w/ integrated WRB", WALL_SFG, "SF", 1.05, 0.55, ""),
        Line("DIV-06", "Stair package (1 flight, ~15 risers)", 1, "EA", 850, 1450, "2-story; pro built"),
        Line("DIV-06", "Interior handrail @ stair", 14, "LF", 8, 18, ""),
        Line("DIV-06", "1x4 trim stock (corners, openings)", 420, "LF", 1.15, 1.85, ""),
        Line("DIV-06", "1x6 head trim", 110, "LF", 1.35, 1.65, ""),
        # DIV-07 THERMAL / MOISTURE
        Line("DIV-07", "Architectural asphalt shingles (main roof)", ROOF_ASPHALT_SF, "SF", 1.65, 2.35,
             "Main roof only — not porch"),
        Line("DIV-07", "Standing-seam metal roof (porch awnings only)", ROOF_METAL_SF, "SF", 6.5, 5.5,
             "Per A2.3 — estimator wrongly applied to full roof"),
        Line("DIV-07", "Synthetic underlayment + ice & water valleys", ROOF_ASPHALT_SF, "SF", 0.42, 0.28, ""),
        Line("DIV-07", "Ridge vent", 52, "LF", 4.5, 3.5, ""),
        Line("DIV-07", "Step / valley flashing", 85, "LF", 3.2, 4.8, ""),
        Line("DIV-07", "6\" lap siding w/ 6\" reveal", WALL_SFG, "SF", 2.45, 3.85, ""),
        Line("DIV-07", "Housewrap / weather barrier", WALL_SFG, "SF", 0.28, 0.18, ""),
        Line("DIV-07", "Soffit w/ venting", 180, "LF", 4.2, 3.8, ""),
        Line("DIV-07", "1x fascia boards", 180, "LF", 2.1, 2.4, ""),
        Line("DIV-07", "R-38 attic insulation", ROOF_ASPHALT_SF, "SF", 1.05, 0.65, ""),
        Line("DIV-07", "R-15/R-20 wall batt (verify w/ 2x4 cavity)", WALL_SFG, "SF", 0.62, 0.38,
             "Plan notes R-20; confirm w/ energy calc"),
        Line("DIV-07", "Sealant at openings", 320, "LF", 0.35, 0.45, ""),
        # DIV-08 OPENINGS — counts from A2.2 door/window schedule
        Line("DIV-08", "3068 full glass entry w/ transom", 1, "EA", 950, 250, ""),
        Line("DIV-08", "3068 half glass exterior door", 2, "EA", 420, 185, ""),
        Line("DIV-08", "2868 interior doors", 3, "EA", 165, 95, ""),
        Line("DIV-08", "2668 interior doors", 8, "EA", 145, 85, ""),
        Line("DIV-08", "2440 closet door", 1, "EA", 95, 65, ""),
        Line("DIV-08", "8-0 x 7-0 garage door (power)", 1, "EA", 1150, 450, ""),
        Line("DIV-08", "32x66 single-hung windows", 5, "EA", 285, 125, ""),
        Line("DIV-08", "64x66 double single-hung windows", 3, "EA", 520, 165, ""),
        Line("DIV-08", "36x60 egress windows (2nd fl)", 2, "EA", 310, 135, "Emergency escape"),
        Line("DIV-08", "Door hardware allowance", 14, "EA", 35, 25, ""),
        # DIV-09 FINISHES
        Line("DIV-09", "1/2\" gypsum board hang + finish (pro tape/mud)", DRYWALL_SF, "SF", 0.42, 1.35,
             "Volunteers hang; pro finish only"),
        Line("DIV-09", "Interior latex paint (walls/ceilings)", DRYWALL_SF, "SF", 0.18, 0.32,
             "Volunteer paint under lead"),
        Line("DIV-09", "LVP flooring — main level living areas", 780, "SF", 2.85, 1.65, ""),
        Line("DIV-09", "Tile flooring — baths + laundry", 145, "SF", 4.5, 8.5, ""),
        Line("DIV-09", "Carpet — bedrooms (2nd floor)", 380, "SF", 2.2, 1.1, ""),
        Line("DIV-09", "4\" rubber wall base", 520, "LF", 1.05, 0.85, ""),
        Line("DIV-09", "Interior trim install (base/case)", 620, "LF", 1.25, 2.15, ""),
        Line("DIV-09", "Exterior paint on siding (as req.)", WALL_SFG, "SF", 0.22, 0.38, ""),
        # DIV-10 SPECIALTIES — residential bath only
        Line("DIV-10", "Bath accessories (bars, TP holder, hooks)", 2, "BATH", 185, 125,
             "Removed commercial dispensers from RW estimate"),
        Line("DIV-10", "Mirrors", 2, "EA", 65, 45, ""),
        # DIV-11 EQUIPMENT — Habitat often donated; budget placeholders
        Line("DIV-11", "Appliance allowance (range, DW, WH, W/D)", 1, "LS", 3200, 450,
             "Donation credit applied at summary"),
        Line("DIV-12", "Kitchen cabinets (10 LF base + 8 LF upper)", 18, "LF", 145, 85, ""),
        Line("DIV-12", "Laminate countertops + 4\" backsplash", 28, "LF", 38, 42,
             "Habitat spec — not solid surface"),
        Line("DIV-12", "Closet shelving / rods (allowance)", 1, "LS", 650, 450, ""),
        # DIV-22 PLUMBING
        Line("DIV-22", "Plumbing rough + trim (3BR/2BA + laundry)", HEATED_SF, "SF", 4.8, 5.2, ""),
        Line("DIV-22", "Water heater (electric 50 gal)", 1, "EA", 650, 450, ""),
        Line("DIV-22", "Plumbing fixtures (WC, lavs, tub, sinks)", 1, "LS", 1850, 950, ""),
        # DIV-23 HVAC
        Line("DIV-23", "Heat pump system (1425 SF conditioned)", 1, "LS", 7200, 3800,
             "Attic air handler per plan note"),
        Line("DIV-23", "Bath exhaust fans (recessed)", 2, "EA", 145, 125, ""),
        Line("DIV-23", "Manual J / duct layout by MECH sub", 1, "LS", 450, 350, ""),
        # DIV-26 ELECTRICAL
        Line("DIV-26", "Electrical rough + trim (2024 IRC)", HEATED_SF, "SF", 5.2, 4.8, ""),
        Line("DIV-26", "Light fixtures (per RCP A2.4)", 1, "LS", 1450, 850, ""),
        Line("DIV-26", "Ceiling fans w/ light", 2, "EA", 165, 95, ""),
        Line("DIV-26", "200A service / panel", 1, "EA", 1850, 950, ""),
        # DIV-31 EARTHWORK
        Line("DIV-31", "Clear, cut/fill, compaction", 85, "CY", 0, 42, "Site-specific"),
        Line("DIV-31", "Soil export / haul-off (allowance)", 45, "CY", 0, 38, ""),
        Line("DIV-31", "Final grade + swale", 1, "LS", 350, 1850, ""),
        # DIV-32 EXTERIOR
        Line("DIV-32", "Driveway (concrete or gravel — allowance)", 420, "SF", 4.5, 3.5, ""),
        Line("DIV-32", "Walks + stoop", 120, "SF", 5.5, 4.5, ""),
        Line("DIV-32", "Sod / seed + landscape allowance", 1, "LS", 1800, 1200, ""),
        # DIV-33 UTILITIES
        Line("DIV-33", "Water/sewer taps + meter fees (allowance)", 1, "LS", 4500, 1500,
             "Confirm w/ Augusta Utilities"),
        Line("DIV-33", "Underground electric service (allowance)", 1, "LS", 2800, 1200, ""),
    ]


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "608 Macon Estimate"

    header_fill = PatternFill("solid", fgColor="1C2833")
    header_font = Font(color="F5F0E8", bold=True)
    div_fill = PatternFill("solid", fgColor="E8E4DC")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:P1")
    ws["A1"] = "608 MACON AVE — HABITAT 3BR ESTIMATE (CORRECTED)"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A2:P2")
    ws["A2"] = (
        f"Source: Booker+Vick permit set A2.2 (May 2026) · "
        f"{HEATED_SF} SF heated · Prepared {date.today().isoformat()}"
    )
    ws["A3"] = "Notes: Habitat Augusta pricing; 10% material waste; volunteer-assisted labor reductions on select trades."

    headers = [
        "DIV", "DESCRIPTION", "QTY", "WASTE", "QTY+WASTE", "UNIT",
        "HAB MAT", "HAB LAB", "TOTAL MAT", "TOTAL LAB", "LINE TOTAL", "NOTES",
        "MKT MAT", "MKT LAB",
    ]
    start = 5
    for col, h in enumerate(headers, 1):
        cell = ws.cell(start, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border

    row = start + 1
    div_totals: dict[str, tuple[float, float]] = {}
    first_data = row

    for item in lines():
        qty_w = q(item.qty) if item.unit not in ("LS", "EA", "WK", "MO", "BATH") else item.qty
        if item.unit in ("LS", "EA", "WK", "MO", "BATH"):
            qty_w = item.qty
        else:
            qty_w = round(item.qty * (1 + WASTE), 2)

        ws.cell(row, 1, item.div).border = border
        ws.cell(row, 2, item.desc).border = border
        ws.cell(row, 3, item.qty).border = border
        ws.cell(row, 4, WASTE if item.unit not in ("LS", "EA", "WK", "MO", "BATH") else 0).border = border
        ws.cell(row, 5, f"=C{row}*(1+D{row})" if item.unit not in ("LS", "EA", "WK", "MO", "BATH") else item.qty).border = border
        ws.cell(row, 6, item.unit).border = border
        ws.cell(row, 7, habitat_mat(item)).border = border
        ws.cell(row, 8, habitat_lab(item)).border = border
        ws.cell(row, 9, f"=E{row}*G{row}").border = border
        ws.cell(row, 10, f"=E{row}*H{row}").border = border
        ws.cell(row, 11, f"=I{row}+J{row}").border = border
        ws.cell(row, 12, item.notes).border = border
        ws.cell(row, 13, item.mat).border = border
        ws.cell(row, 14, item.lab).border = border

        div_totals.setdefault(item.div, [0.0, 0.0])
        row += 1

    last_data = row - 1

    # Summary block
    row += 1
    ws.cell(row, 2, "SUBTOTAL (direct costs)").font = Font(bold=True)
    ws.cell(row, 9, f"=SUM(I{first_data}:I{last_data})")
    ws.cell(row, 10, f"=SUM(J{first_data}:J{last_data})")
    ws.cell(row, 11, f"=SUM(K{first_data}:K{last_data})")
    sub_row = row

    row += 1
    ws.cell(row, 2, f"Contingency ({CONTINGENCY:.0%})")
    ws.cell(row, 11, f"=K{sub_row}*{CONTINGENCY}")
    cont_row = row

    row += 1
    ws.cell(row, 2, f"Overhead ({OH_RATE:.0%})")
    ws.cell(row, 11, f"=(K{sub_row}+K{cont_row})*{OH_RATE}")
    oh_row = row

    row += 1
    ws.cell(row, 2, "Habitat donation credits (appliances/materials)")
    ws.cell(row, 11, -DONATION_CREDIT)
    don_row = row

    row += 1
    ws.cell(row, 2, "TOTAL PROJECT BUDGET").font = Font(bold=True, size=12)
    ws.cell(row, 11, f"=K{sub_row}+K{cont_row}+K{oh_row}+K{don_row}").font = Font(bold=True, size=12)
    total_row = row

    row += 2
    ws.cell(row, 2, "Cost per heated SF (Habitat program)").font = Font(bold=True)
    ws.cell(row, 11, f"=K{total_row}/{HEATED_SF}")
    ws.cell(row, 12, "Target band: $95–$115/SF Habitat Augusta")

    row += 1
    ws.cell(row, 2, "Market-rate GC equivalent (reference)")
    ws.cell(row, 11, "=SUMPRODUCT(E{0}:E{1},M{0}:M{1})+SUMPRODUCT(E{0}:E{1},N{0}:N{1})".format(first_data, last_data))
    ws.cell(row, 12, "What RW-style commercial pricing would yield")

    # Assumptions sheet
    ass = wb.create_sheet("Assumptions")
    assumptions = [
        ("Project", "608 Macon Avenue, Augusta GA 30901"),
        ("Program", "Habitat for Humanity — 3 BR / 2 BA"),
        ("Architect", "Booker + Vick (Job 2615, May 2026)"),
        ("Heated SF", str(HEATED_SF)),
        ("First floor SF", str(FIRST_FLOOR_SF)),
        ("Second floor SF", str(SECOND_FLOOR_SF)),
        ("Garage SF", str(GARAGE_SF)),
        ("Covered porch SF", str(PORCH_SF)),
        ("", ""),
        ("Corrections from RW estimator", ""),
        ("Areas", "RW used 1094 slab / 1305 MEP — permit set is 1425 heated SF"),
        ("Foundation", "RW used 15x14 strip footings — permit is thickened slab-on-grade"),
        ("Roof", "RW doubled asphalt + metal on full roof — metal is porch-only per A2.3"),
        ("Framing", "RW used 2x4 floor joists — plan specifies wood I-joists 2nd floor"),
        ("Finishes", "RW solid surface counters + commercial bath accessories removed"),
        ("HVAC", "RW missing system — heat pump added for 1425 SF"),
        ("Utilities", "RW DIV-33 at $0 — tap fees added as allowance"),
        ("Demolition", "RW DIV-02 included — not in scope for this new build"),
        ("Donation credits", f"${DONATION_CREDIT:,} applied for typical Habitat in-kind support"),
    ]
    for i, (k, v) in enumerate(assumptions, 1):
        ass.cell(i, 1, k).font = Font(bold=True)
        ass.cell(i, 2, v)

    # Column widths
    widths = [10, 52, 8, 8, 10, 8, 10, 10, 12, 12, 12, 40, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ass.column_dimensions["A"].width = 28
    ass.column_dimensions["B"].width = 72

    # Number formats
    for r in range(first_data, total_row + 1):
        for c in [3, 4, 5, 7, 8, 9, 10, 11]:
            ws.cell(r, c).number_format = '#,##0.00'

    cmp = wb.create_sheet("RW vs Corrected")
    cmp["A1"] = "RW Estimator Issue"
    cmp["B1"] = "Corrected (608 Macon permit set)"
    for cell in [cmp["A1"], cmp["B1"]]:
        cell.font = Font(bold=True)
    comparisons = [
        ("Heated square footage", "RW: ~1,305 SF → Correct: 1,425 SF (A2.2)"),
        ("Slab area", "RW: 1,094 SF → Correct: ~1,324 SF (952 + 252 garage + porch)"),
        ("Foundation", "RW: 15×14 strip footings → Correct: thickened slab-on-grade"),
        ("Roof", "RW: asphalt + metal full roof → Correct: asphalt main + metal porch only"),
        ("Framing", "RW: 2x4 joists + 2x8 rafters → Correct: trusses + I-joists"),
        ("HVAC", "RW: missing → Correct: heat pump 1,425 SF"),
        ("Utilities", "RW: $0 → Correct: tap fee allowance"),
        ("Pricing", "RW: $0 total (empty) → Correct: Habitat + market reference columns"),
    ]
    for i, (a, b) in enumerate(comparisons, 2):
        cmp.cell(i, 1, a)
        cmp.cell(i, 2, b)
    cmp.column_dimensions["A"].width = 24
    cmp.column_dimensions["B"].width = 72

    return wb


def main() -> None:
    wb = build_workbook()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    wb.save(DOWNLOADS)
    print(f"Wrote {OUT}")
    print(f"Wrote {DOWNLOADS}")


if __name__ == "__main__":
    main()
